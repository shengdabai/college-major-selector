#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_admission_json.py — 把「本省历年录取数据」转换成本工具可上传的 JSON 模板。

本脚本是【通用模板】，适用于任意省份、任意科类（物理类 / 历史类 / 文科 / 理科 / 3+3 选科组合）。
它本身不内置任何省份数据；你需要：
  1) 从本省教育考试院 / 阳光高考平台获取公开的「招生计划 / 专业录取分 / 院校录取分 / 一分一段表」Excel；
  2) 按下方 COLUMN_MAP 把你 Excel 的列名映射到本工具需要的字段；
  3) 运行：python3 scripts/build_admission_json.py <你的Excel.xlsx> <年份> --province 河南 --subject 物理类
     生成 data/<年份>.json，再在网页「03 录取数据」里上传即可。

输出 JSON 的字段契约（与 app.js 解析逻辑一致）：
{
  "_meta":  { "province": "河南", "year": 2025, "subject": "物理类", "plans": N, "majorScores": N },
  "plans":        [ { "school","sCode","major","mCode","mCat","mGroup","note",
                       "subject","req","type","batch","plan","years","fee","lang" } ],
  "majorScores":  [ { "school","major","mCode","batch","mGroup",
                       "minScore","minRank","is985","is211","owner","province","note" } ],
  "schoolScores": [ { "school","batch","type","minScore","minRank","lineDiff",
                       "admitted","is985","is211","owner","province" } ],
  "rankTable":    [ { "score", "cum", "line" } ]   # 一分一段表：分数 -> 累计人数，line=本批次控制线
}

字段说明：
  minScore  专业/院校最低投档分          minRank  对应最低位次
  req       选科要求（如「物理+化学」「不限」）  batch    批次（如「本科批」）
  plan      招生计划人数                  fee      学费（元/年，字符串）
  rankTable 一分一段表，用于「分数 ↔ 位次」换算，是分数推荐的基础

提示：没有编程基础也可以直接在网页上点「下载 JSON 模板/示例」，照着模板手工填几条数据即可使用。
"""

import json
import os
import sys
import argparse

try:
    from openpyxl import load_workbook
except ImportError:
    print("需要 openpyxl：请先运行  pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(ROOT, "data")

# ============================================================
# 按你自己的 Excel 列名修改这里（左边是本工具字段，右边是你表头里的列名）
# 不同省份 / 不同年份的官方表格列名不一样，这里务必按实际表头调整。
# ============================================================
COLUMN_MAP = {
    "school": "院校名称",
    "sCode": "院校代号",
    "major": "专业名称",
    "mCode": "专业代号",
    "batch": "批次",
    "req": "选科要求",
    "plan": "计划数",
    "fee": "学费",
    "minScore": "最低分",
    "minRank": "最低位次",
}


def cell(v):
    return str(v).strip() if v is not None else ""


def to_int(v, default=0):
    try:
        return int(float(str(v).replace(",", "").strip()))
    except (ValueError, TypeError):
        return default


def read_rows(xlsx_path):
    """读取 Excel 第一个 sheet，返回 [dict]，key 为表头列名。"""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [cell(h) for h in rows[0]]
    out = []
    for r in rows[1:]:
        if r is None:
            continue
        rec = {header[i]: r[i] for i in range(min(len(header), len(r)))}
        if any(cell(v) for v in rec.values()):
            out.append(rec)
    return out


def col(rec, field):
    """按 COLUMN_MAP 取值。"""
    src = COLUMN_MAP.get(field)
    return rec.get(src) if src else None


def build(xlsx_path, year, province, subject):
    rows = read_rows(xlsx_path)
    plans, major_scores = [], []
    for rec in rows:
        school = cell(col(rec, "school"))
        major = cell(col(rec, "major"))
        if not school or not major:
            continue
        batch = cell(col(rec, "batch")) or "本科批"
        plans.append({
            "school": school, "sCode": cell(col(rec, "sCode")),
            "major": major, "mCode": cell(col(rec, "mCode")),
            "mCat": "", "mGroup": "", "note": "",
            "subject": subject, "req": cell(col(rec, "req")) or "不限",
            "type": "普通类", "batch": batch,
            "plan": to_int(col(rec, "plan")), "years": "四年",
            "fee": cell(col(rec, "fee")), "lang": "不限",
        })
        min_score = to_int(col(rec, "minScore"), 0)
        min_rank = to_int(col(rec, "minRank"), 0)
        if min_score or min_rank:
            major_scores.append({
                "school": school, "major": major, "mCode": cell(col(rec, "mCode")),
                "batch": batch, "mGroup": "",
                "minScore": min_score, "minRank": min_rank,
                "is985": False, "is211": False, "owner": "公办",
                "province": "", "note": "",
            })

    result = {
        "_meta": {
            "province": province, "year": int(year), "subject": subject,
            "plans": len(plans), "majorScores": len(major_scores),
        },
        "plans": plans,
        "majorScores": major_scores,
        "schoolScores": [],   # 如有院校层面投档数据，按字段契约自行补充
        "rankTable": [],      # 一分一段表请单独整理后填入：[{score, cum, line}]
    }
    return result


def main():
    ap = argparse.ArgumentParser(description="把本省录取 Excel 转成本工具可上传的 JSON")
    ap.add_argument("xlsx", help="招生计划 / 录取分 Excel 路径")
    ap.add_argument("year", help="数据年份，如 2025")
    ap.add_argument("--province", default="本省", help="省份名，如 河南")
    ap.add_argument("--subject", default="物理类", help="科类，如 物理类 / 历史类 / 文科 / 理科")
    ap.add_argument("--out", default=None, help="输出路径，默认 data/<年份>.json")
    args = ap.parse_args()

    data = build(args.xlsx, args.year, args.province, args.subject)
    os.makedirs(DATA_DIR, exist_ok=True)
    out = args.out or os.path.join(DATA_DIR, f"{args.year}.json")
    with open(out, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"已生成 {out}")
    print(f"  plans={data['_meta']['plans']}  majorScores={data['_meta']['majorScores']}")
    print("提示：rankTable（一分一段表）需自行整理后填入，否则分数↔位次换算无法使用。")


if __name__ == "__main__":
    main()
