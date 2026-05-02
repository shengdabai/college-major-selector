#!/usr/bin/env python3
"""解析青海 22-25 历年招生计划/录取分数/一分一段表 -> data/qh_*.json
仅保留 物理类/理科。"""
import json, os, re
from collections import defaultdict
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, '青海22-25数据')
DATA = os.path.join(ROOT, 'data')
os.makedirs(DATA, exist_ok=True)

# 物理类 / 理科 视为同一阵营
def is_physics_track(s):
    s = (s or '').strip()
    return s in ('物理类', '理科')

def s(v):
    return str(v).strip() if v is not None else ''

def i(v, default=0):
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return default

# ---------- 1) 招生计划（按 院校+专业+年份+科类 索引） ----------
def parse_plans():
    """52110 -> 仅物理类约 1/2"""
    wb = load_workbook(os.path.join(SRC, '22-25年全国高校在青海的招生计划.xlsx'),
                       read_only=True, data_only=True)
    ws = wb.worksheets[0]
    out = []
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx == 0: continue
        if not row[0]: continue
        # 0年份 1院校 2院校代码 3科类 4批次 5招生类型 6专业 7专业代码
        # 8专业组 9备注 10选科 11招生人数 12学制 13学费
        if not is_physics_track(row[3]): continue
        out.append({
            'year': i(row[0]),
            'school': s(row[1]),
            'sCode': s(row[2]),
            'track': s(row[3]),
            'batch': s(row[4]),
            'type': s(row[5]) or '普通类',
            'major': s(row[6]),
            'mCode': s(row[7]),
            'mGroup': s(row[8]),
            'note': s(row[9]),
            'req': s(row[10]) or '不限',
            'plan': i(row[11]),
            'years': s(row[12]),
            'fee': s(row[13]),
        })
    return out

# ---------- 2) 专业录取分数 ----------
def parse_major_scores():
    wb = load_workbook(os.path.join(SRC, '22-25年全国高校在青海的专业录取分数.xlsx'),
                       read_only=True, data_only=True)
    ws = wb.worksheets[0]
    out = []
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx == 0: continue
        if not row[0]: continue
        # 0年份 1院校 2院校代码 3科类 4批次 5专业 6专业代码 7专业组 8备注
        # 9选科 10录取人数 11最低分数 12最低位次 13学校所在 14学校性质 15是否985 16是否211
        if not is_physics_track(row[3]): continue
        out.append({
            'year': i(row[0]),
            'school': s(row[1]),
            'sCode': s(row[2]),
            'track': s(row[3]),
            'batch': s(row[4]),
            'major': s(row[5]),
            'mCode': s(row[6]),
            'mGroup': s(row[7]),
            'note': s(row[8]),
            'req': s(row[9]) or '不限',
            'admitted': i(row[10]),
            'minScore': i(row[11]),
            'minRank': i(row[12]),
            'province': s(row[13]),
            'owner': s(row[14]),
            'is985': s(row[15]) == '是',
            'is211': s(row[16]) == '是',
        })
    return out

# ---------- 3) 院校录取分数 ----------
def parse_school_scores():
    wb = load_workbook(os.path.join(SRC, '22-25年全国高校在青海的院校录取分数.xlsx'),
                       read_only=True, data_only=True)
    ws = wb.worksheets[0]
    out = []
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx == 0: continue
        if not row[0]: continue
        # 0年份 1院校 2院校代码 3科类 4批次 5招生类型 6专业组 7选科
        # 8录取人数 9最低分数 10最低分位 11批次线差 12学校所在 13学校性质 14是否985 15是否211
        if not is_physics_track(row[3]): continue
        out.append({
            'year': i(row[0]),
            'school': s(row[1]),
            'sCode': s(row[2]),
            'track': s(row[3]),
            'batch': s(row[4]),
            'type': s(row[5]) or '普通类',
            'mGroup': s(row[6]),
            'req': s(row[7]) or '不限',
            'admitted': i(row[8]),
            'minScore': i(row[9]),
            'minRank': i(row[10]),
            'lineDiff': i(row[11]),  # 批次线差
            'province': s(row[12]),
            'owner': s(row[13]),
            'is985': s(row[14]) == '是',
            'is211': s(row[15]) == '是',
        })
    return out

# ---------- 4) 一分一段表 ----------
def parse_score_rank():
    """{ year: [ {score, count, cumulative, rangeStart, rangeEnd, line, batch} ] }"""
    out = {}
    for y in [2022, 2023, 2024, 2025]:
        wb = load_workbook(os.path.join(SRC, f'青海{y}年的一分一段表.xlsx'),
                           read_only=True, data_only=True)
        ws = wb.worksheets[0]
        rows = []
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            if idx == 0: continue
            if not row[0]: continue
            track = s(row[1])
            if not is_physics_track(track): continue
            score_str = s(row[4])  # "632-750" 或 "621"
            count = i(row[5])
            cumulative = i(row[6])
            line = i(row[3])
            batch = s(row[2])
            # 取分数下界（用于分数→位次查询）
            m = re.match(r'^(\d+)', score_str)
            score_low = int(m.group(1)) if m else 0
            # 上界
            m2 = re.search(r'-(\d+)$', score_str)
            score_high = int(m2.group(1)) if m2 else score_low
            rows.append({
                'score': score_low,
                'scoreHigh': score_high,
                'count': count,
                'cum': cumulative,    # 累计人数 = 位次
                'line': line,         # 控制线
                'batch': batch,
            })
        # 按分数降序
        rows.sort(key=lambda r: -r['score'])
        out[y] = rows
        print(f'  一分一段 {y}: {len(rows)} 行（控制线 {rows[0]["line"] if rows else "?"}）')
    return out

if __name__ == '__main__':
    print('解析招生计划…')
    plans = parse_plans()
    print(f'  → {len(plans)} 条')

    print('解析专业录取分数…')
    major_scores = parse_major_scores()
    print(f'  → {len(major_scores)} 条')

    print('解析院校录取分数…')
    school_scores = parse_school_scores()
    print(f'  → {len(school_scores)} 条')

    print('解析一分一段表…')
    rank_table = parse_score_rank()

    # 拆分按年份输出，前端可按需加载
    by_year = defaultdict(lambda: {'plans': [], 'majorScores': [], 'schoolScores': []})
    for p in plans: by_year[p['year']]['plans'].append(p)
    for r in major_scores: by_year[r['year']]['majorScores'].append(r)
    for r in school_scores: by_year[r['year']]['schoolScores'].append(r)

    for year in [2022, 2023, 2024, 2025]:
        d = by_year[year]
        d['year'] = year
        d['province'] = '青海'
        d['track'] = '物理类/理科'
        d['rankTable'] = rank_table.get(year, [])
        d['_meta'] = {
            'plans': len(d['plans']),
            'majorScores': len(d['majorScores']),
            'schoolScores': len(d['schoolScores']),
            'rankRows': len(d['rankTable']),
        }
        path = os.path.join(DATA, f'qh_{year}.json')
        json.dump(d, open(path, 'w', encoding='utf-8'),
                  ensure_ascii=False, separators=(',', ':'))
        size = os.path.getsize(path) / 1024
        print(f'  {path} {size:.1f} KB  {d["_meta"]}')

    # 2026 占位
    p26 = os.path.join(DATA, 'qh_2026.json')
    if not os.path.exists(p26):
        json.dump({
            'year': 2026,
            'province': '青海',
            'track': '物理类/理科',
            'plans': [], 'majorScores': [], 'schoolScores': [], 'rankTable': [],
            '_placeholder': True,
            '_note': '2026 数据公布后运行 scripts/parse_qinghai_history.py 后替换此文件',
        }, open(p26, 'w', encoding='utf-8'),
        ensure_ascii=False, indent=2)
        print(f'  + 占位 {p26}')

    # 删除旧 qinghai_*.json（被新文件取代）
    for old in ['qinghai_2025.json', 'qinghai_2026.json']:
        p = os.path.join(DATA, old)
        if os.path.exists(p):
            os.remove(p)
            print(f'  - 删除旧文件 {old}')
