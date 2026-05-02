#!/usr/bin/env python3
"""解析青海招生计划 xlsx -> data/qinghai_<year>.json
紧凑列式存储以减小体积。"""
import json, os, sys, re
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, 'data')
os.makedirs(DATA, exist_ok=True)

def parse(xlsx_path, year):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb['青海'] if '青海' in wb.sheetnames else wb.worksheets[0]
    plans = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2: continue
        if row[0] is None or row[2] is None: continue
        # 0省份 1年份 2院校 3院校代码 4专业 5专业代码 6专业类 7专业备注
        # 8科目 9选科要求 10类型 11批次 12计划人数 13学制 14学费 15外语
        plans.append({
            'school': row[2],
            'sCode': str(row[3]) if row[3] is not None else '',
            'major': row[4],
            'mCode': str(row[5]) if row[5] is not None else '',
            'mCat': row[6] or '',
            'note': row[7] or '',
            'subject': row[8] or '',          # 物理/历史
            'req': row[9] or '',               # 选科要求
            'type': row[10] or '',             # 预科/民族班/藏文班/...
            'batch': row[11] or '',            # 本科批/专科批/...
            'plan': int(row[12]) if isinstance(row[12], (int, float)) else 0,
            'years': int(row[13]) if isinstance(row[13], (int, float)) else 0,
            'fee': str(row[14]) if row[14] is not None else '',
            'lang': row[15] or '不限',
        })
    return plans

def build(plans, year):
    schools = sorted({p['school'] for p in plans})
    majors = sorted({p['major'] for p in plans})
    out = {
        'year': year,
        'province': '青海',
        'total': len(plans),
        'schools': schools,
        'majors': majors,
        'plans': plans,
    }
    return out

if __name__ == '__main__':
    src = sys.argv[1] if len(sys.argv) > 1 else os.path.join(ROOT, '20250627-青海2025招生计划.xlsx')
    year = int(sys.argv[2]) if len(sys.argv) > 2 else 2025
    plans = parse(src, year)
    data = build(plans, year)
    out_path = os.path.join(DATA, f'qinghai_{year}.json')
    json.dump(data, open(out_path, 'w', encoding='utf-8'),
              ensure_ascii=False, separators=(',', ':'))
    size = os.path.getsize(out_path)
    print(f'{out_path}  {size/1024:.1f} KB  {len(plans)} 条计划  {len(data["schools"])} 所院校  {len(data["majors"])} 个专业')

    # 创建 2026 占位文件（首次运行时；后续不覆盖）
    p2026 = os.path.join(DATA, 'qinghai_2026.json')
    if not os.path.exists(p2026):
        json.dump({
            'year': 2026,
            'province': '青海',
            'total': 0,
            'schools': [],
            'majors': [],
            'plans': [],
            '_placeholder': True,
            '_note': '等 2026 招生计划公布后，运行 scripts/parse_qinghai.py <xlsx> 2026 替换此文件',
        }, open(p2026, 'w', encoding='utf-8'),
        ensure_ascii=False, indent=2)
        print(f'  + 创建占位 {p2026}')
